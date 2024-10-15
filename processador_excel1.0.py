import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

def selecionar_arquivo():
    # Abre uma janela para o usuário selecionar o arquivo
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal do Tkinter
    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    return file_path

def gerar_lista_aprovados(aprovados):
    # Perguntar ao usuário se deseja gerar a lista de aprovados
    resposta = input("Você gostaria de salvar a lista de aprovados em um arquivo .txt? (s/n): ")
    
    if resposta.lower() == 's':
        with open("aprovados.txt", "w") as f:
            for item in aprovados:
                f.write(f"{item}\n")
        print("Lista de aprovados gerada em 'aprovados.txt'.")
    else:
        print("Lista de aprovados não foi salva.")

def analisar_primeira_coluna(file_path):
    # Ler o arquivo Excel
    df = pd.read_excel(file_path)
    
    # Selecionar a primeira coluna e remover duplicatas
    primeira_coluna = df.iloc[:, 0].drop_duplicates()
    
    # Criar listas para aprovados e erros com base nos critérios
    aprovados = []
    erros = []
    linhas_de_erros = []

    for index, num in enumerate(primeira_coluna):
        # Converter o valor para string e remover hifens e espaços
        num_limpo = re.sub(r'\D', '', str(num))
        
        # Verificar se o número limpo começa com "6" e tem comprimento múltiplo de 7
        if num_limpo.startswith("6") and len(num_limpo) % 7 == 0:
            # Dividir o número em partes de 7 dígitos e adicionar cada parte em 'aprovados'
            partes = [num_limpo[i:i+7] for i in range(0, len(num_limpo), 7)]
            aprovados.extend(partes)
        else:
            erros.append(num_limpo)
            linhas_de_erros.append(index + 2)  # Adiciona 2 para considerar o cabeçalho e a indexação

    # Criar um novo arquivo Excel para salvar as alterações
    novo_workbook = Workbook()
    novo_sheet = novo_workbook.active
    
    # Adicionar cabeçalho
    for col in range(1, df.shape[1] + 1):
        novo_sheet.cell(row=1, column=col, value=df.columns[col - 1])
    
    # Adicionar os dados e colorir linhas de erro
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            novo_sheet.cell(row=index + 2, column=col_index + 1, value=value)
        
        # Se a linha for de erro, pintar de amarelo
        if index + 2 in linhas_de_erros:
            for cell in novo_sheet[index + 2]:
                cell.fill = yellow_fill

    # Salvar as alterações em um novo arquivo Excel
    novo_workbook.save("erros.xlsx")

    # Exibir os resultados dos aprovados e erros
    print("Aprovados:", aprovados)
    print("Erros:", erros)
    print("As linhas com erros foram pintadas de amarelo no arquivo 'arquivo_modificado.xlsx'.")

    # Chamar a função para gerar a lista de aprovados
    gerar_lista_aprovados(aprovados)

# Executa o seletor de arquivos e passa o arquivo selecionado para a função
file_path = selecionar_arquivo()
if file_path:
    analisar_primeira_coluna(file_path)
else:
    print("Nenhum arquivo foi selecionado.")
