import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re
from tkinter import messagebox

def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw()  
    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    return file_path

def gerar_lista_aprovados(aprovados, exibir_mensagem):
    resposta = messagebox.askyesno("Salvar lista", "Você gostaria de salvar a lista de aprovados em um arquivo .txt?")
    if resposta:
        with open("numeros.txt", "w") as f:
            for item in aprovados:
                f.write(f"{item}\n")
        exibir_mensagem("Lista de aprovados gerada em\n'numeros.txt'.")
    else:
        exibir_mensagem("Lista de aprovados não foi salva.")

def analisar_primeira_coluna(file_path, exibir_mensagem=None):
    df = pd.read_excel(file_path)
    primeira_coluna = df.iloc[:, 0].drop_duplicates()
    
    aprovados = []
    erros = []
    linhas_de_erros = []

    for index, num in primeira_coluna.items():
        num_limpo = re.sub(r'\D', '', str(num))
        if num_limpo.startswith("6") and len(num_limpo) % 7 == 0:
            partes = [num_limpo[i:i+7] for i in range(0, len(num_limpo), 7)]
            aprovados.extend(partes)
        else:
            erros.append(num_limpo)
            linhas_de_erros.append(index)

    novo_workbook = Workbook()
    novo_sheet = novo_workbook.active
    
    for col in range(1, df.shape[1] + 1):
        novo_sheet.cell(row=1, column=col, value=df.columns[col - 1])
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            novo_sheet.cell(row=index + 2, column=col_index + 1, value=value)
        
        if index in linhas_de_erros:
            for cell in novo_sheet[index + 2]:
                cell.fill = yellow_fill

    novo_workbook.save("erros.xlsx")


    if exibir_mensagem:
        exibir_mensagem(f"Total de erros: {len(erros)}")
        exibir_mensagem(f"Total de aprovados: {len(aprovados)}")
        exibir_mensagem("As linhas com erros foram pintadas de amarelo no arquivo 'erros.xlsx'.")

    gerar_lista_aprovados(aprovados, exibir_mensagem)

def main():
    root = tk.Tk()
    root.title("Processador de Excel")

    label_mensagem = tk.Label(root, text="Clique para processar o arquivo", font=("Arial", 14))
    label_mensagem.pack(pady=20)
    
    def atualizar_mensagem(nova_mensagem):
        label_mensagem.config(text=nova_mensagem)

    def processar():
        file_path = selecionar_arquivo()
        if file_path:
            atualizar_mensagem("Processando o arquivo...")
            analisar_primeira_coluna(file_path, atualizar_mensagem)
        else:
            atualizar_mensagem("Nenhum arquivo foi selecionado.")

    botao = tk.Button(root, text="Iniciar Processamento", command=processar)
    botao.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
